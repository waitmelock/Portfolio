#-*- coding: utf-8 -*-
from PyQt5 import QtWidgets, QtCore, QtGui
from SCJSC_ui import Ui_MainWindow
from pop_out_ui import Ui_Form
#import pymysql
import openpyxl
import os
import glob

dataset = {'time':"",'applicant':"",'class':"",'amount':0, 'item':"", 'comment':"", 'activity':"", 'group':""}
dataset_budget_dance = {'class':"舞風費",'FF':0,'P':0,'B':0,'L':0,'H':0,'total':0}
dataset_budget_ide = {'class':"期初舞展",'攝影組':0,'美宣組':0,'器材組':0,'total':0} #init dance exhibition
dataset_budget_mde = {'class':"小成",'攝影組':0,'美宣組':0,'器材組':0,'total':0} #mini dance exhibition
dataset_budget_fde = {'class':"大成",'攝影組':0,'美宣組':0,'器材組':0,'total':0} #final dance exhibition
dataset_dance_expenditure = {'class':"舞風費",'FF':0,'P':0,'B':0,'L':0,'H':0,'total':0}
dataset_ide_expenditure = {'class':"期初舞展",'攝影組':0,'美宣組':0,'器材組':0,'total':0}
dataset_mde_expenditure = {'class':"小成",'攝影組':0,'美宣組':0,'器材組':0,'total':0}
dataset_fde_expenditure = {'class':"大成",'攝影組':0,'美宣組':0,'器材組':0,'total':0}


#creat worksheets in excel
pathname=r".\SCJSC.xlsx"
#本來沒有這個檔案
if glob.glob(pathname)==[]:
    workbook=openpyxl.Workbook()
    journal_account = workbook.worksheets[0]
    income_account=workbook.create_sheet('收入')
    workbook.create_sheet('支出')
    workbook.create_sheet('預算')
    journal_account.title='所有收支'
    income_account=workbook.worksheets[1]
    expenditure_account=workbook.worksheets[2]
    budget_account=workbook.worksheets[3]
    
    dataset_number = 1 #列數
    dataset_keys_number = 0 #行數


    for key in dataset.keys():
        journal_account.cell(row=dataset_number, column=dataset_keys_number+1).value = key
        dataset_keys_number += 1
    dataset_number += 1

    dataset_number = 1 #列數
    dataset_keys_number = 0 #行數

    dataset_keys_number = 0 
    for key in dataset_budget_dance.keys():
        budget_account.cell(row=dataset_number, column=dataset_keys_number+1).value = key
        budget_account.cell(row=dataset_number+1, column=dataset_keys_number+1).value = dataset_budget_dance[key]
        expenditure_account.cell(row=dataset_number, column=dataset_keys_number+1).value = key
        expenditure_account.cell(row=dataset_number+1, column=dataset_keys_number+1).value = dataset_dance_expenditure[key]
        dataset_keys_number += 1
    dataset_number += 2

    dataset_keys_number = 0    
    for key in dataset_budget_ide.keys():
        budget_account.cell(row=dataset_number, column=dataset_keys_number+1).value = key
        budget_account.cell(row=dataset_number+1, column=dataset_keys_number+1).value = dataset_budget_ide[key]
        expenditure_account.cell(row=dataset_number, column=dataset_keys_number+1).value = key
        expenditure_account.cell(row=dataset_number+1, column=dataset_keys_number+1).value = dataset_ide_expenditure[key]
        dataset_keys_number += 1
    dataset_number += 2

    dataset_keys_number = 0 
    for key in dataset_budget_mde.keys():
        budget_account.cell(row=dataset_number, column=dataset_keys_number+1).value = key
        budget_account.cell(row=dataset_number+1, column=dataset_keys_number+1).value = dataset_budget_mde[key]
        expenditure_account.cell(row=dataset_number, column=dataset_keys_number+1).value = key
        expenditure_account.cell(row=dataset_number+1, column=dataset_keys_number+1).value = dataset_mde_expenditure[key]
        dataset_keys_number += 1
    dataset_number += 2
    
    dataset_keys_number = 0 
    for key in dataset_budget_fde.keys():
        budget_account.cell(row=dataset_number, column=dataset_keys_number+1).value = key
        budget_account.cell(row=dataset_number+1, column=dataset_keys_number+1).value = dataset_budget_fde[key]
        expenditure_account.cell(row=dataset_number, column=dataset_keys_number+1).value = key
        expenditure_account.cell(row=dataset_number+1, column=dataset_keys_number+1).value = dataset_fde_expenditure[key]
        dataset_keys_number += 1
    dataset_number += 2



    workbook.save('SCJSC.xlsx') 
#本來就有這個檔案   
else:
    workbook = openpyxl.load_workbook('SCJSC.xlsx')
    journal_account = workbook['所有收支']
    income_account=workbook['收入']
    expenditure_account=workbook['支出']
    budget_account=workbook['預算']
    workbook.save('SCJSC.xlsx')

class Main(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):

        self.password = "0000"
        self.data = dataset
        self.budget_dance=dataset_budget_dance
        self.budget_ide=dataset_budget_ide
        self.budget_mde=dataset_budget_mde
        self.budget_fde=dataset_budget_fde
        self.dance_expenditure = dataset_dance_expenditure
        self.ide_expenditure = dataset_ide_expenditure
        self.mde_expenditure = dataset_mde_expenditure
        self.fde_expenditure = dataset_fde_expenditure


        set_target = '舞風費'
        for numRow in range(1,budget_account.max_row,1):
            produceName = budget_account.cell(row=numRow, column=1).value
            if produceName == set_target:
                self.budget_dance['FF'] = budget_account.cell(row=numRow, column=2).value
                self.dance_expenditure['FF'] = expenditure_account.cell(row=numRow, column=2).value
                self.budget_dance['P'] = budget_account.cell(row=numRow, column=3).value
                self.dance_expenditure['P'] = expenditure_account.cell(row=numRow, column=3).value
                self.budget_dance['B'] = budget_account.cell(row=numRow, column=4).value
                self.dance_expenditure['B'] = expenditure_account.cell(row=numRow, column=4).value
                self.budget_dance['L'] = budget_account.cell(row=numRow, column=5).value
                self.dance_expenditure['L'] = expenditure_account.cell(row=numRow, column=5).value
                self.budget_dance['H'] = budget_account.cell(row=numRow, column=6).value
                self.dance_expenditure['H'] = expenditure_account.cell(row=numRow, column=6).value
                self.budget_dance['total'] = budget_account.cell(row=numRow, column=7).value
                self.dance_expenditure['total'] = expenditure_account.cell(row=numRow, column=7).value

        set_target = '期初舞展'
        for numRow in range(1,budget_account.max_row,1):
            produceName = budget_account.cell(row=numRow, column=1).value
            if produceName == set_target:
                self.budget_ide['攝影組'] = budget_account.cell(row=numRow, column=2).value
                self.ide_expenditure['攝影組'] = expenditure_account.cell(row=numRow, column=2).value
                self.budget_ide['美宣組'] = budget_account.cell(row=numRow, column=3).value
                self.ide_expenditure['美宣組'] = expenditure_account.cell(row=numRow, column=3).value
                self.budget_ide['器材組'] = budget_account.cell(row=numRow, column=4).value
                self.ide_expenditure['器材組'] = expenditure_account.cell(row=numRow, column=4).value
                self.budget_ide['total'] = budget_account.cell(row=numRow, column=5).value
                self.ide_expenditure['total'] = expenditure_account.cell(row=numRow, column=5).value

        set_target = '小成'
        for numRow in range(1,budget_account.max_row,1):
            produceName = budget_account.cell(row=numRow, column=1).value
            if produceName == set_target:
                self.budget_mde['攝影組'] = budget_account.cell(row=numRow, column=2).value
                self.mde_expenditure['攝影組'] = expenditure_account.cell(row=numRow, column=2).value
                self.budget_mde['美宣組'] = budget_account.cell(row=numRow, column=3).value
                self.mde_expenditure['美宣組'] = expenditure_account.cell(row=numRow, column=3).value
                self.budget_mde['器材組'] = budget_account.cell(row=numRow, column=4).value
                self.mde_expenditure['器材組'] = expenditure_account.cell(row=numRow, column=4).value
                self.budget_mde['total'] = budget_account.cell(row=numRow, column=5).value
                self.mde_expenditure['total'] = expenditure_account.cell(row=numRow, column=5).value

        set_target = '大成'
        for numRow in range(1,budget_account.max_row+1,1):
            produceName = budget_account.cell(row=numRow, column=1).value
            if produceName == set_target:
                self.budget_fde['攝影組'] = budget_account.cell(row=numRow, column=2).value
                self.fde_expenditure['攝影組'] = expenditure_account.cell(row=numRow, column=2).value
                self.budget_fde['美宣組'] = budget_account.cell(row=numRow, column=3).value
                self.fde_expenditure['美宣組'] = expenditure_account.cell(row=numRow, column=3).value
                self.budget_fde['器材組'] = budget_account.cell(row=numRow, column=4).value
                self.fde_expenditure['器材組'] = expenditure_account.cell(row=numRow, column=4).value
                self.budget_fde['total'] = budget_account.cell(row=numRow, column=5).value
                self.fde_expenditure['total'] = expenditure_account.cell(row=numRow, column=5).value




        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.setup_control()
        
    def setup_control(self):
        self.ui.comboBox_expenditure_activity.activated.connect(lambda:self.comBox_to_label(self.ui.label_expenditure_con_activity_2, self.ui.comboBox_expenditure_activity.currentText()))   #當combox activity改變時 用comBox_to_label 把combo裡的字顯示在label上
        self.ui.comboBox_expenditure_activity.activated.connect(lambda:self.com_act_gp_set(self.ui.comboBox_expenditure_activity,self.ui.comboBox_expenditure_group))                       #當combox改變時 用com_act_gp_set 把activity對應的group設定好
        self.ui.comboBox_expenditure_activity.activated.connect(lambda:self.comBox_to_label(self.ui.label_expenditure_con_group_2,self.ui.comboBox_expenditure_group.currentText()))
        self.ui.comboBox_expenditure_activity.activated.connect(lambda:self.set_act_prograss_val())
        self.ui.comboBox_expenditure_activity.activated.connect(lambda:self.set_gp_prograss_val())
        self.ui.comboBox_expenditure_group.activated.connect(lambda:self.comBox_to_label(self.ui.label_expenditure_con_group_2, self.ui.comboBox_expenditure_group.currentText()))            #當combox group改變時 用comBox_to_label 把combo裡的字顯示在label上
        self.ui.comboBox_expenditure_group.activated.connect(lambda:self.set_gp_prograss_val())
        self.ui.pushButton_expenditure_check.clicked.connect(lambda:self.receive_expenditure_data())                                                                                        #當pushButton_expenditure_check按下 用 receive_expenditure_data 接收填寫的資料
        self.ui.pushButton_expenditure_check.clicked.connect(lambda:self.set_act_prograss_val())
        self.ui.pushButton_expenditure_check.clicked.connect(lambda:self.set_gp_prograss_val())                                                                                                #更新活動預算prograssbar
        self.ui.pushButton_income_check.clicked.connect(lambda:self.receive_income_data())                                                                                                  #當pushButton_income_check按下 用 receive_income_data 接收填寫的資料
        self.ui.dateEdit_expenditure.setDate(QtCore.QDate().currentDate())  #設定支出的初始日期(開啟程式的日期)
        self.ui.dateEdit__income.setDate(QtCore.QDate().currentDate())      #設定收入的初始日期(開啟程式的日期)
        self.ui.progressBar_expenditure_con_activity.setValue(0)            #設定進度條初始
        self.ui.progressBar_expenditure_con_group.setValue(0)               #設定進度條初始

        self.pop_window = PopWindow() #宣告彈出視窗的object
        self.ui.lineEdit_condition_password_2.returnPressed.connect(lambda :self.password_check(self.ui.lineEdit_condition_password_2.text())) #當lineEdit_condition_password_2按下enter後 用password_check檢查密碼
        
        self.pop_window.ui.pushButton_check.clicked.connect(lambda:self.receive_budget())
        self.ui.comboBox_setting_activity.activated.connect(lambda:self.budget_show())
        self.pop_window.ui.comboBox_activity.activated.connect(lambda:self.com_act_gp_set(self.pop_window.ui.comboBox_activity,self.pop_window.ui.comboBox_group))
        



    #comBox_to_label把combo裡的text顯示在label上
    def comBox_to_label(self, label, text):
        label.setText(text)

    # 接收支出填寫的資料
    def receive_expenditure_data(self):
        if self.ui.lineEdit_expenditure_people.text() == "" or self.ui.lineEdit_expenditure_money.text()=="" or self.ui.lineEdit_expenditure_name.text()=="" or self.ui.textEdit_expenditure_remark.toPlainText()=="":
            return self.ui.label_expenditure_foruser.setText("資料輸入錯誤") #為了避免程式crash要阻止空的line edit去assign給其他變數
        else:
            self.ui.label_expenditure_foruser.setText("資料儲存成功")
        self.data['class'] = "支出" 
        self.data['applicant'] = self.ui.lineEdit_expenditure_people.text()         #申請人
        self.data['amount'] = int(self.ui.lineEdit_expenditure_money.text())        #金額
        self.data['item'] = self.ui.lineEdit_expenditure_name.text()                #項目
        self.data['comment'] = self.ui.textEdit_expenditure_remark.toPlainText()     #備註
        self.data['time'] = self.ui.dateEdit_expenditure.date().toString()          #時間
        self.data['activity'] = self.ui.comboBox_expenditure_activity.currentText() #活動
        self.data['group'] = self.ui.comboBox_expenditure_group.currentText()       #組別
          
        if self.data['activity'] == '舞風費':
            if self.data['group'] == 'FF':
                self.dance_expenditure['FF'] += self.data['amount']
            elif self.data['group'] == 'P':
                self.dance_expenditure['P'] += self.data['amount']
            elif self.data['group'] == 'B':
                self.dance_expenditure['B'] += self.data['amount']
            elif self.data['group'] == 'L':
                self.dance_expenditure['L'] += self.data['amount']
            elif self.data['group'] == 'H':
                self.dance_expenditure['H'] += self.data['amount']
        elif self.data['activity'] == '期初舞展':
            if self.data['group'] == '攝影組':
                self.ide_expenditure['攝影組'] += self.data['amount']
            elif self.data['group'] == '美宣組':
                self.ide_expenditure['美宣組'] += self.data['amount']
            elif self.data['group'] == '器材組':
                self.ide_expenditure['器材組'] += self.data['amount']
        elif self.data['activity'] == '小成':
            if self.data['group'] == '攝影組':
                self.mde_expenditure['攝影組'] += self.data['amount']
            elif self.data['group'] == '美宣組':
                self.mde_expenditure['美宣組'] += self.data['amount']
            elif self.data['group'] == '器材組':
                self.mde_expenditure['器材組'] += self.data['amount']
        elif self.data['activity'] == '大成':
            if self.data['group'] == '攝影組':
                self.fde_expenditure['攝影組'] += self.data['amount']
            elif self.data['group'] == '美宣組':
                self.fde_expenditure['美宣組'] += self.data['amount']
            elif self.data['group'] == '器材組':
                self.fde_expenditure['器材組'] += self.data['amount']
        
        self.dance_expenditure['total'] = self.dance_expenditure['FF'] + self.dance_expenditure['P'] + self.dance_expenditure['B'] + self.dance_expenditure['L'] + self.dance_expenditure['H']
        self.ide_expenditure['total'] = self.ide_expenditure['攝影組'] + self.ide_expenditure['美宣組'] + self.ide_expenditure['器材組']
        self.mde_expenditure['total'] = self.mde_expenditure['攝影組'] + self.mde_expenditure['美宣組'] + self.mde_expenditure['器材組']
        self.fde_expenditure['total'] = self.fde_expenditure['攝影組'] + self.fde_expenditure['美宣組'] + self.fde_expenditure['器材組']
        self.xl_record()#記錄資料進excel  

    # 接收收入填寫的資料       
    def receive_income_data(self):
        if self.ui.lineEdit_income_people.text() == "" or self.ui.lineEdit_income_money.text()=="" or self.ui.lineEdit_income_name.text()=="" or self.ui.textEdit_income_remark.toPlainText()=="":
            return self.ui.label_income_foruser.setText("資料輸入錯誤")#為了避免程式crash要阻止空的line edit去assign給其他變數
        else:
            self.ui.label_income_foruser.setText("資料儲存成功")
        self.data['class'] = "收入" 
        self.data['applicant'] = self.ui.lineEdit_income_people.text()         #申請人
        self.data['amount'] = int(self.ui.lineEdit_income_money.text())        #金額
        self.data['item'] = self.ui.lineEdit_income_name.text()                #項目
        self.data['comment'] = self.ui.textEdit_income_remark.toPlainText()    #備註
        self.data['time'] = self.ui.dateEdit__income.date().toString()         #時間
        self.data['activity'] = self.ui.comboBox_income_activity.currentText() #活動
        self.data['group'] = self.ui.comboBox__income_class.currentText()      #組別
        self.xl_record()#記錄資料進excel    

    # 檢查密碼
    def password_check(self, password):
        if self.password == password:
            self.pop_window.show()
        else:
            self.pop_window.hide()
        
    #記錄資料進excel    
    def xl_record(self):
        self.data_number=journal_account.max_row
        self.data_keys_number = 0
        for value in self.data.values():
            journal_account.cell(row=self.data_number+1, column=self.data_keys_number+1).value = value
            self.data_keys_number += 1


        set_target = '舞風費'
        for numRow in range(1,budget_account.max_row,1):
            produceName = budget_account.cell(row=numRow, column=1).value
            if produceName == set_target:
                budget_account.cell(row=numRow, column=2).value = self.budget_dance['FF']
                expenditure_account.cell(row=numRow, column=2).value = self.dance_expenditure['FF']
                budget_account.cell(row=numRow, column=3).value = self.budget_dance['P']
                expenditure_account.cell(row=numRow, column=3).value = self.dance_expenditure['P']
                budget_account.cell(row=numRow, column=4).value = self.budget_dance['B']
                expenditure_account.cell(row=numRow, column=4).value = self.dance_expenditure['B']
                budget_account.cell(row=numRow, column=5).value = self.budget_dance['L']
                expenditure_account.cell(row=numRow, column=5).value = self.dance_expenditure['L']
                budget_account.cell(row=numRow, column=6).value = self.budget_dance['H']
                expenditure_account.cell(row=numRow, column=6).value = self.dance_expenditure['H']
                budget_account.cell(row=numRow, column=7).value = self.budget_dance['total']
                expenditure_account.cell(row=numRow, column=7).value = self.dance_expenditure['total']

        set_target = '期初舞展'
        for numRow in range(1,budget_account.max_row,1):
            produceName = budget_account.cell(row=numRow, column=1).value
            if produceName == set_target:
                budget_account.cell(row=numRow, column=2).value = self.budget_ide['攝影組']
                expenditure_account.cell(row=numRow, column=2).value = self.ide_expenditure['攝影組']
                budget_account.cell(row=numRow, column=3).value = self.budget_ide['美宣組']
                expenditure_account.cell(row=numRow, column=3).value = self.ide_expenditure['美宣組']
                budget_account.cell(row=numRow, column=4).value = self.budget_ide['器材組']
                expenditure_account.cell(row=numRow, column=4).value = self.ide_expenditure['器材組']
                budget_account.cell(row=numRow, column=5).value = self.budget_ide['total']
                expenditure_account.cell(row=numRow, column=5).value = self.ide_expenditure['total']

        set_target = '小成'
        for numRow in range(1,budget_account.max_row,1):
            produceName = budget_account.cell(row=numRow, column=1).value
            if produceName == set_target:
                budget_account.cell(row=numRow, column=2).value = self.budget_mde['攝影組']
                expenditure_account.cell(row=numRow, column=2).value = self.mde_expenditure['攝影組']
                budget_account.cell(row=numRow, column=3).value = self.budget_mde['美宣組']
                expenditure_account.cell(row=numRow, column=3).value = self.mde_expenditure['美宣組']
                budget_account.cell(row=numRow, column=4).value = self.budget_mde['器材組']
                expenditure_account.cell(row=numRow, column=4).value = self.mde_expenditure['器材組']
                budget_account.cell(row=numRow, column=5).value = self.budget_mde['total']
                expenditure_account.cell(row=numRow, column=5).value = self.mde_expenditure['total']

        set_target = '大成'
        for numRow in range(1,budget_account.max_row+1,1):
            produceName = budget_account.cell(row=numRow, column=1).value
            if produceName == set_target:
                budget_account.cell(row=numRow, column=2).value = self.budget_fde['攝影組']
                expenditure_account.cell(row=numRow, column=2).value = self.fde_expenditure['攝影組']
                budget_account.cell(row=numRow, column=3).value = self.budget_fde['美宣組']
                expenditure_account.cell(row=numRow, column=3).value = self.fde_expenditure['美宣組']
                budget_account.cell(row=numRow, column=4).value = self.budget_fde['器材組']
                expenditure_account.cell(row=numRow, column=4).value = self.fde_expenditure['器材組']
                budget_account.cell(row=numRow, column=5).value = self.budget_fde['total']
                expenditure_account.cell(row=numRow, column=5).value = self.fde_expenditure['total']

        
        workbook.save('SCJSC.xlsx')

    def xl_budget_record(self):
        self.dataset_budget_dance_number=2
        self.dataset_budget_dance_keys_number = 0
        for value in self.budget_dance.values():
            budget_account.cell(row=self.dataset_budget_dance_number, column=self.dataset_budget_dance_keys_number+1).value = value
            self.dataset_budget_dance_keys_number += 1
        self.dataset_budget_dance_number += 2

        self.dataset_budget_dance_keys_number=0
        for value in self.budget_ide.values():
            budget_account.cell(row=self.dataset_budget_dance_number, column=self.dataset_budget_dance_keys_number+1).value = value
            self.dataset_budget_dance_keys_number += 1
        self.dataset_budget_dance_number += 2

        self.dataset_budget_dance_keys_number=0
        for value in self.budget_mde.values():
            budget_account.cell(row=self.dataset_budget_dance_number, column=self.dataset_budget_dance_keys_number+1).value = value
            self.dataset_budget_dance_keys_number += 1
        self.dataset_budget_dance_number += 2
        
        self.dataset_budget_dance_keys_number=0
        for value in self.budget_fde.values():
            budget_account.cell(row=self.dataset_budget_dance_number, column=self.dataset_budget_dance_keys_number+1).value = value
            self.dataset_budget_dance_keys_number += 1
        self.dataset_budget_dance_number += 2

        workbook.save('SCJSC.xlsx')
        
    def popwin_com_set(self):
        self.pop_window.ui.comboBox_group.clear()
        if self.pop_window.ui.comboBox_activity.currentText() == "舞風費":
            self.pop_window.ui.comboBox_group.addItem("FF")
            self.pop_window.ui.comboBox_group.addItem("P")
            self.pop_window.ui.comboBox_group.addItem("B")
            self.pop_window.ui.comboBox_group.addItem("L")
            self.pop_window.ui.comboBox_group.addItem("H")
        else:
            self.pop_window.ui.comboBox_group.addItem("攝影組")
            self.pop_window.ui.comboBox_group.addItem("美宣組")
            self.pop_window.ui.comboBox_group.addItem("器材組")

    #把combox的activity對應的group設定好
    def com_act_gp_set(self,UI_combox_act,UI_combox_gp):
        UI_combox_gp.clear()
        if UI_combox_act.currentText() == "舞風費":
            UI_combox_gp.addItem("FF")
            UI_combox_gp.addItem("P")
            UI_combox_gp.addItem("B")
            UI_combox_gp.addItem("L")
            UI_combox_gp.addItem("H")
        else :
            UI_combox_gp.addItem("攝影組")
            UI_combox_gp.addItem("美宣組")
            UI_combox_gp.addItem("器材組")
            if UI_combox_act != self.pop_window.ui.comboBox_activity: #如果不是popwindow就多一欄個人
                UI_combox_gp.addItem("個人")

    #接收預算
    def receive_budget(self):
        if self.pop_window.ui.comboBox_activity.currentText() == self.budget_dance['class']:
            self.budget_dance[self.pop_window.ui.comboBox_group.currentText()] = int(self.pop_window.ui.lineEdit.text())
        elif self.pop_window.ui.comboBox_activity.currentText() == self.budget_ide['class']:
            self.budget_ide[self.pop_window.ui.comboBox_group.currentText()] = int(self.pop_window.ui.lineEdit.text())    
        elif self.pop_window.ui.comboBox_activity.currentText() == self.budget_mde['class']:
            self.budget_mde[self.pop_window.ui.comboBox_group.currentText()] = int(self.pop_window.ui.lineEdit.text())
        elif self.pop_window.ui.comboBox_activity.currentText() == self.budget_fde['class']:
            self.budget_fde[self.pop_window.ui.comboBox_group.currentText()] = int(self.pop_window.ui.lineEdit.text())
        
        self.pop_window.ui.label_budget_showcondition.setText("輸入成功")
        self.cal_total_budge()
        self.budget_show()
        self.xl_budget_record()
        self.set_act_prograss_val()
        self.set_gp_prograss_val()

    #顯示預算
    def budget_show(self):
        if self.ui.comboBox_setting_activity.currentText() !="請選擇項目":
            self.ui.label_setting_1.show()
            self.ui.label_setting_money_1.show()
            self.ui.label_setting_2.show()
            self.ui.label_setting_money_2.show()
            self.ui.label_setting_3.show()
            self.ui.label_setting_money_3.show()
            self.ui.label_setting_4.show()
            self.ui.label_setting_money_4.show()
            self.ui.label_setting_5.show()
            self.ui.label_setting_money_5.show()
            if self.ui.comboBox_setting_activity.currentText() == self.budget_dance['class']:
                self.ui.label_setting_1.setText("FF")
                self.ui.label_setting_money_1.setText(str(self.budget_dance['FF']))
                self.ui.label_setting_2.setText("P")
                self.ui.label_setting_money_2.setText(str(self.budget_dance['P']))
                self.ui.label_setting_3.setText("B")
                self.ui.label_setting_money_3.setText(str(self.budget_dance['B']))
                self.ui.label_setting_4.setText("L")
                self.ui.label_setting_money_4.setText(str(self.budget_dance['L']))
                self.ui.label_setting_5.setText("H")
                self.ui.label_setting_money_5.setText(str(self.budget_dance['H']))
            elif self.ui.comboBox_setting_activity.currentText() == self.budget_ide['class']:
                self.budget_set_label_de(self.budget_ide)
            elif self.ui.comboBox_setting_activity.currentText() == self.budget_mde['class']:
                self.budget_set_label_de(self.budget_mde)
            elif self.ui.comboBox_setting_activity.currentText() == self.budget_fde['class']:
                self.budget_set_label_de(self.budget_fde)
        else:
            self.ui.label_setting_1.hide()
            self.ui.label_setting_money_1.hide()
            self.ui.label_setting_2.hide()
            self.ui.label_setting_money_2.hide()
            self.ui.label_setting_3.hide()
            self.ui.label_setting_money_3.hide()
            self.ui.label_setting_4.hide()
            self.ui.label_setting_money_4.hide()
            self.ui.label_setting_5.hide()
            self.ui.label_setting_money_5.hide()
            
    def budget_set_label_de(self,budget_dic):
        self.ui.label_setting_1.setText('攝影組')
        self.ui.label_setting_money_1.setText(str(budget_dic['攝影組']))
        self.ui.label_setting_2.setText('美宣組')
        self.ui.label_setting_money_2.setText(str(budget_dic['美宣組']))
        self.ui.label_setting_3.setText('器材組')
        self.ui.label_setting_money_3.setText(str(budget_dic['器材組']))
        self.ui.label_setting_4.hide()
        self.ui.label_setting_money_4.hide()
        self.ui.label_setting_5.hide()
        self.ui.label_setting_money_5.hide()
    #計算個活動總預算
    def cal_total_budge(self):
        self.budget_dance['total'] = self.budget_dance['FF'] + self.budget_dance['P'] + self.budget_dance['B'] + self.budget_dance['L'] + self.budget_dance['H']
        self.budget_ide['total'] = self.budget_ide['攝影組'] + self.budget_ide['美宣組'] + self.budget_ide['器材組']
        self.budget_mde['total'] = self.budget_mde['攝影組'] + self.budget_mde['美宣組'] + self.budget_mde['器材組']
        self.budget_fde['total'] = self.budget_fde['攝影組'] + self.budget_fde['美宣組'] + self.budget_fde['器材組']
    
    def set_act_prograss_val(self):
        if self.ui.label_expenditure_con_activity_2.text() == '舞風費' and self.budget_dance['total'] != 0:
            self.ui.progressBar_expenditure_con_activity.setValue( int((self.dance_expenditure['total'] * 100) / self.budget_dance['total']))
        elif self.ui.label_expenditure_con_activity_2.text() == '期初舞展' and self.budget_ide['total'] != 0:
            self.ui.progressBar_expenditure_con_activity.setValue( int((self.ide_expenditure['total'] * 100) / self.budget_ide['total']))
        elif self.ui.label_expenditure_con_activity_2.text() == '小成' and self.budget_mde['total'] != 0:
            self.ui.progressBar_expenditure_con_activity.setValue( int((self.mde_expenditure['total'] * 100) / self.budget_mde['total']))
        elif self.ui.label_expenditure_con_activity_2.text() == '大成' and self.budget_fde['total'] != 0:
            self.ui.progressBar_expenditure_con_activity.setValue( int((self.fde_expenditure['total'] * 100) / self.budget_fde['total']))
        else:
            self.ui.progressBar_expenditure_con_activity.setValue(0)
    
    def set_gp_prograss_val(self):
        text = self.ui.label_expenditure_con_group_2.text()
        if self.ui.label_expenditure_con_activity_2.text() == '舞風費' and self.budget_dance[text] != 0:
            self.ui.progressBar_expenditure_con_group.setValue( int((self.dance_expenditure[text] * 100) / self.budget_dance[text]))
        elif self.ui.label_expenditure_con_activity_2.text() == '期初舞展' and self.budget_ide[text] != 0:
            self.ui.progressBar_expenditure_con_group.setValue( int((self.ide_expenditure[text] * 100) / self.budget_ide[text]))
        elif self.ui.label_expenditure_con_activity_2.text() == '小成' and self.budget_mde[text] != 0:
            self.ui.progressBar_expenditure_con_group.setValue( int((self.mde_expenditure[text] * 100) / self.budget_mde[text]))
        elif self.ui.label_expenditure_con_activity_2.text() == '大成' and self.budget_fde[text] != 0:
            self.ui.progressBar_expenditure_con_group.setValue( int((self.fde_expenditure[text] * 100) / self.budget_fde[text]))
        else:
            self.ui.progressBar_expenditure_con_group.setValue(0)
    

        
    
class PopWindow(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)
    

        
'''
    #管理者登入與註冊

    def register(self, email, password, name):
        print("Register")

    def login(self,email,password):
        for person in self.manager:
            if person[0]==email:
                if person[1]==password:
                    print("Good")
                else:
                    print("Not Good")
            else:
                print("Not exist")
'''